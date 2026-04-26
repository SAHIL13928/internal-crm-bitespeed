"""
Enrich the Master sheet of Fireflies Mapping by filling missing email
addresses for shopUrls using every available signal:

  A. Meet-link <> shopUrl join (Arindam CSV × Fireflies scrape external attendees)
  B. Emails extracted directly from Arindam message content
  C. Group-name brand → meeting-title contains → external attendees
  D. ShopUrl <> Brand Name sheet → meeting-title contains → external attendees
  E. Fuzzy Brand Name Match sheet (tier >= 1) → meeting → external attendees
  F. Shop-slug → meeting-title contains → external attendees
"""
import csv
import re
import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict

MASTER_XLSX_IN = "Fireflies Mapping (1).xlsx"
MASTER_XLSX_OUT = "Fireflies Mapping (1) - ENRICHED.xlsx"
SCRAPE_CSV = "all_meet_links_organized_v2.csv"
ARINDAM_CSV = "meetlinkstoshopUrl (1).csv"

MEET_RE = re.compile(r"https://meet\.google\.com/[a-z0-9\-]+", re.IGNORECASE)
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

INTERNAL_DOMAINS = {"bitespeed.co"}
JUNK_DOMAINS = {
    "gmail.com", "yahoo.com", "outlook.com", "hotmail.com", "icloud.com",
    "googlemail.com", "proton.me", "protonmail.com", "live.com", "aol.com",
}


def is_internal(email):
    return email.lower().split("@")[-1] in INTERNAL_DOMAINS


def parse_emails_from_attendee_str(s):
    """Pulls emails out of 'name <email>; name2 <email2>' or 'email; email' format."""
    out = set()
    if not s:
        return out
    for part in str(s).split("; "):
        part = part.strip()
        m = re.search(r"<(.+?)>", part)
        if m:
            email = m.group(1).strip().lower()
        elif "@" in part:
            email = part.strip().lower()
        else:
            continue
        if email and not is_internal(email):
            out.add(email)
    return out


def normalize_brand(s):
    """Aggressive brand slug for fuzzy contains matching."""
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"\.myshopify\.com$", "", s)
    s = re.sub(r"\.com$|\.in$|\.co$|\.store$|\.shop$", "", s)
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


STOPWORDS = {
    # Common, non-distinguishing words that match hundreds of meetings
    "india", "store", "shop", "shoppe", "online", "world", "global", "group",
    "brand", "brands", "inc", "ltd", "pvt", "company", "co", "labs", "studio",
    "official", "original", "style", "styles", "fashion", "fashions",
    "collection", "collections", "kids", "women", "woman", "mens", "beauty",
    "home", "decor", "living", "goods", "luxe", "luxury", "premium", "best",
    "bitespeed", "bite", "speed", "call", "meet", "meeting", "sync", "weekly",
    "biweekly", "biw", "kickoff", "kick", "connect", "review", "intro",
    "onboarding", "demo", "chat", "sales", "aicallback", "ai", "voice",
    "support", "team", "wa", "whatsapp", "organic", "natural", "healthy",
    "care", "love", "life", "first", "new", "the", "pro", "plus", "max",
    "app", "club", "store", "foods", "food", "drink", "drinks", "wellness",
    "cafe", "biotech", "nutrition", "hair", "skin", "body", "baby", "mom",
    "cosmetic", "cosmetics", "jewellery", "jewelry", "clothing", "apparel",
    "wear", "essentials", "basics", "shirts", "tshirt", "tshirts", "pant",
    "pants", "store", "outlet", "mart", "bazaar", "enterprises",
}


def brand_tokens(s, max_tokens=2):
    """Significant distinguishing tokens. Returns at most the top `max_tokens`
    longest, non-stopword tokens. Prefer the LONGEST token as most distinctive."""
    if not s:
        return []
    s = str(s).lower()
    s = re.sub(r"\.myshopify\.com$", "", s)
    s = re.sub(r"\<\>.*$", "", s)
    parts = re.split(r"[^a-z0-9]+", s)
    parts = [p for p in parts if len(p) >= 4 and p not in STOPWORDS]
    # Rank by length desc, keep top max_tokens
    parts.sort(key=lambda p: (-len(p), p))
    return parts[:max_tokens]


# Precompute how many meetings a given token would match. Tokens that match
# too many meetings are generic — we reject them at match time.
def _count_title_matches(token):
    pat = re.compile(r"\b" + re.escape(token) + r"\b", re.IGNORECASE)
    return sum(1 for r in scrape_rows if pat.search(r["title"]))


_token_freq_cache = {}


def token_ok(token):
    """Reject tokens that appear in too many meeting titles (generic)."""
    if token in _token_freq_cache:
        return _token_freq_cache[token]
    c = _count_title_matches(token)
    ok = c <= 25
    _token_freq_cache[token] = ok
    return ok


def match_titles(tokens):
    """Return scrape rows whose title word-matches the distinctive tokens.

    If 2+ tokens: AND-match (title must contain every token).
    If 1 token: single match, but only if the token is distinctive enough
                (<= 6 matches, dropped if generic) and length >= 5.
    """
    tokens = [t for t in tokens if token_ok(t)]
    if not tokens:
        return []
    patterns = [re.compile(r"\b" + re.escape(t) + r"\b", re.IGNORECASE) for t in tokens]
    if len(patterns) == 1:
        # Single-token matching is dangerous — require stricter threshold.
        if len(tokens[0]) < 5:
            return []
        return [
            r for r in scrape_rows
            if r["external_emails"] and patterns[0].search(r["title"])
        ]
    # AND-match: every token must appear somewhere in the title
    return [
        r for r in scrape_rows
        if r["external_emails"] and all(p.search(r["title"]) for p in patterns)
    ]


# ── Load meeting scrape and build lookups ────────────────────────────────
scrape_rows = []
with open(SCRAPE_CSV, encoding="utf-8") as f:
    for r in csv.DictReader(f):
        scrape_rows.append({
            "meeting_id": r.get("meeting_id", ""),
            "title": (r.get("meeting_title") or "").strip(),
            "meeting_link": (r.get("meeting_link") or "").strip().lower(),
            "external_emails": parse_emails_from_attendee_str(r.get("external_attendees") or ""),
            "title_norm": normalize_brand(r.get("meeting_title") or ""),
        })

title_to_emails = defaultdict(set)
link_to_emails = defaultdict(set)
for r in scrape_rows:
    if r["meeting_link"]:
        link_to_emails[r["meeting_link"]] |= r["external_emails"]
    if r["title"]:
        title_to_emails[r["title"].lower()] |= r["external_emails"]


# ── Source A: Arindam meet link -> shopUrl (majority vote) joined with scrape ─
link_shop_counts = defaultdict(lambda: defaultdict(int))
link_group = {}
shop_groupnames = defaultdict(set)
content_rows_per_shop = defaultdict(list)
emails_from_content = defaultdict(set)

with open(ARINDAM_CSV, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        content = row.get("content") or ""
        shop = (row.get("shopUrl") or "").strip().lower()
        group = (row.get("group_name") or "").strip()
        if not shop:
            continue

        if group:
            shop_groupnames[shop].add(group)

        # Meet links
        for link in MEET_RE.findall(content):
            link_shop_counts[link.lower()][shop] += 1
            link_group[link.lower()] = group

        # B: emails directly in content (calendar invites shared in the group)
        for em in EMAIL_RE.findall(content):
            em = em.lower()
            if is_internal(em):
                continue
            emails_from_content[shop].add(em)

meet_to_shop = {}
for link, shop_counts in link_shop_counts.items():
    winner = max(shop_counts.items(), key=lambda kv: kv[1])[0]
    meet_to_shop[link] = winner

shop_to_emails_A = defaultdict(set)
for r in scrape_rows:
    ml = r["meeting_link"]
    shop = meet_to_shop.get(ml)
    if shop and r["external_emails"]:
        shop_to_emails_A[shop] |= r["external_emails"]


# ── Load Master + brand sheets ────────────────────────────────────────────
# Two copies: one for reading computed values (data_only), one for writing
wb_vals = openpyxl.load_workbook(MASTER_XLSX_IN, data_only=True)
wb = openpyxl.load_workbook(MASTER_XLSX_IN)
ws_master = wb["Master"]
ws_master_vals = wb_vals["Master"]
ws_brands = wb_vals["ShopUrl <> Brand Name"]
ws_fuzzy = wb_vals["Fuzzy Brand Name Match"]
ws_mlbn = wb_vals["Meeting Link <> Brand Name"]

# ShopUrl -> canonical brand name
shop_to_brand = {}
for i, row in enumerate(ws_brands.iter_rows(values_only=True)):
    if i == 0:
        continue
    name, shop = row[0], row[1]
    if shop and isinstance(shop, str):
        s = shop.strip().lower()
        if name and isinstance(name, str) and s:
            shop_to_brand[s] = name.strip()

# Fuzzy sheet title -> shopUrl (tier >= 1)
fuzzy_title_to_shop = {}
for i, row in enumerate(ws_fuzzy.iter_rows(values_only=True)):
    if i == 0:
        continue
    title, matched_brand, shop, tier = row[1], row[2], row[3], row[4]
    try:
        t = float(tier) if tier is not None else 0
    except (TypeError, ValueError):
        t = 0
    if shop and isinstance(shop, str) and title and t >= 1:
        fuzzy_title_to_shop[str(title).strip().lower()] = str(shop).strip().lower()

# Meeting Link <> Brand Name sheet: authoritative meeting_link -> shopUrl
mlbn_link_to_shop = {}
mlbn_shop_to_emails = defaultdict(set)
for i, row in enumerate(ws_mlbn.iter_rows(values_only=True)):
    if i == 0:
        continue
    ext = row[4] if len(row) > 4 else None
    link = row[3] if len(row) > 3 else None
    shop_h = row[7] if len(row) > 7 else None
    if link and isinstance(link, str) and shop_h and isinstance(shop_h, str):
        s = shop_h.strip().lower()
        if s and s not in ("#n/a", "n/a"):
            mlbn_link_to_shop[link.strip().lower()] = s
            mlbn_shop_to_emails[s] |= parse_emails_from_attendee_str(ext)


# ── Source C: group name -> title contains (strict tokens) ────────────────
shop_to_emails_C = defaultdict(set)
for shop, groups in shop_groupnames.items():
    for g in groups:
        tokens = brand_tokens(g)
        matches = match_titles(tokens)
        for r in matches:
            shop_to_emails_C[shop] |= r["external_emails"]


# ── Source D: ShopUrl <> Brand Name sheet -> title contains (strict) ──────
shop_to_emails_D = defaultdict(set)
for shop, brand in shop_to_brand.items():
    tokens = brand_tokens(brand)
    matches = match_titles(tokens)
    for r in matches:
        shop_to_emails_D[shop] |= r["external_emails"]


# ── Source E: Fuzzy Brand Name Match sheet (tier >= 1) ────────────────────
shop_to_emails_E = defaultdict(set)
for title_l, shop in fuzzy_title_to_shop.items():
    emails = title_to_emails.get(title_l, set())
    if emails:
        shop_to_emails_E[shop] |= emails


# ── Source F: Shop slug -> title contains (strict) ────────────────────────
shop_to_emails_F = defaultdict(set)


def shop_slug_tokens(shop):
    base = shop
    base = re.sub(r"\.myshopify\.com$", "", base)
    base = re.sub(r"\.(com|in|co|store|shop)$", "", base)
    parts = re.split(r"[^a-z0-9]+", base)
    parts = [p for p in parts if len(p) >= 4
             and p not in STOPWORDS
             and not re.fullmatch(r"[0-9a-f]{4,}", p)]
    parts.sort(key=lambda p: (-len(p), p))
    return parts[:2]


def looks_like_hash(shop):
    base = re.sub(r"\.myshopify\.com$", "", shop)
    return bool(re.fullmatch(r"[0-9a-f\-]{6,}", base))


def compute_F_for(shops):
    out = defaultdict(set)
    for shop in shops:
        if looks_like_hash(shop):
            continue
        toks = shop_slug_tokens(shop)
        matches = match_titles(toks)
        for r in matches:
            out[shop] |= r["external_emails"]
    return out


# ── Source G: Global domain-pool scan. For each shop, find emails in the
# entire external pool whose @-domain starts with a distinctive shop token.
# This surfaces attendees from meetings that never got linked to this shop. ─
all_external_emails = set()
for r in scrape_rows:
    all_external_emails |= r["external_emails"]

def domain_base(em):
    dom = em.split("@")[-1].lower() if "@" in em else ""
    return dom.split(".")[0] if dom else ""


def compute_G_for(shops):
    """Global domain scan: match email domain base to the shop's distinctive
    token. Uses ONLY shop-slug-derived tokens plus brand-sheet tokens — NOT
    group-name tokens (too noisy). Requires exact base-label match or
    base-starts-with-token-then-separator, to avoid 'health'→'healthymaster'
    substring hits.
    """
    out = defaultdict(set)
    for shop in shops:
        tokens = set()
        if shop in shop_to_brand:
            tokens.update(brand_tokens(shop_to_brand[shop]))
        tokens.update(shop_slug_tokens(shop))
        tokens = {t for t in tokens if t and len(t) >= 5}
        if not tokens:
            continue
        for em in all_external_emails:
            base = domain_base(em)
            if not base:
                continue
            for tok in tokens:
                if base == tok or base.startswith(tok + "-") or base.startswith(tok + "_"):
                    out[shop].add(em)
                    break
    return out


# ── Source B already built: emails_from_content ───────────────────────────


# ── Walk Master, fill missing ─────────────────────────────────────────────
master_shops = []
header = None
rows_cache = []
# zip the editable sheet with the values-only sheet
val_rows = list(ws_master_vals.iter_rows(values_only=True))
edit_rows = list(ws_master.iter_rows(values_only=False))
for i, (v_row, e_row) in enumerate(zip(val_rows, edit_rows)):
    if i == 0:
        header = list(v_row)
        continue
    shop = v_row[0]
    if not shop:
        continue
    rendered_email = v_row[1]
    s = str(shop).strip().lower()
    master_shops.append(s)
    # (row_index 1-based, shop, editable email cell, rendered email value)
    rows_cache.append((i + 1, s, e_row[1], rendered_email))

# Compute source F and G over full master set
shop_to_emails_F_master = compute_F_for(master_shops)
shop_to_emails_G = compute_G_for(master_shops)

# Combine all sources
def combined_for(shop):
    return (
        shop_to_emails_A.get(shop, set())
        | mlbn_shop_to_emails.get(shop, set())
        | emails_from_content.get(shop, set())
        | shop_to_emails_C.get(shop, set())
        | shop_to_emails_D.get(shop, set())
        | shop_to_emails_E.get(shop, set())
        | shop_to_emails_F.get(shop, set())
        | shop_to_emails_F_master.get(shop, set())
        | shop_to_emails_G.get(shop, set())
    )


def sources_for(shop):
    tags = []
    if shop_to_emails_A.get(shop): tags.append("A")
    if mlbn_shop_to_emails.get(shop): tags.append("A'")
    if emails_from_content.get(shop): tags.append("B")
    if shop_to_emails_C.get(shop): tags.append("C")
    if shop_to_emails_D.get(shop): tags.append("D")
    if shop_to_emails_E.get(shop): tags.append("E")
    if shop_to_emails_F.get(shop) or shop_to_emails_F_master.get(shop): tags.append("F")
    if shop_to_emails_G.get(shop): tags.append("G")
    return tags


# ── Post-filter: score emails, prefer ones whose @-domain matches the shop
# brand, and drop candidates that appear across many different shops (agency
# noise). ─────────────────────────────────────────────────────────────────
email_to_shops = defaultdict(set)
for shop in master_shops:
    cand = combined_for(shop)
    for em in cand:
        email_to_shops[em].add(shop)

# Highly agency-like emails match many shops; cap at 4 (safe).
AGENCY_CAP = 4


def filter_candidates(shop, cand):
    if not cand:
        return cand
    # Determine strong tokens for this shop
    tokens = set()
    if shop in shop_to_brand:
        tokens.update(brand_tokens(shop_to_brand[shop]))
    tokens.update(shop_slug_tokens(shop))
    for g in shop_groupnames.get(shop, set()):
        tokens.update(brand_tokens(g))

    # Which candidates have an @-domain whose base begins with / exactly
    # matches a strong token? Requiring the token at the start of the domain
    # (or as the full left-side of the SLD) blocks substring mis-hits like
    # "bombay"-token matching "bombayshirts.com".
    tokens_sorted = sorted(tokens, key=lambda t: -len(t))

    def domain_matches(em):
        dom = em.split("@")[-1].lower() if "@" in em else ""
        if not dom:
            return False
        base = dom.split(".")[0]
        for tok in tokens_sorted:
            if not tok or len(tok) < 4:
                continue
            if base == tok or base.startswith(tok + "-") or base.startswith(tok + "_"):
                return True
        return False

    domain_matching = {em for em in cand if domain_matches(em)}

    if domain_matching:
        # High-confidence: use only the domain-matching candidates
        return domain_matching

    # No domain match at all — treat as low-confidence. Drop emails that
    # appear as candidates for many unrelated shops (agency/vendor noise).
    cleaned = {em for em in cand if len(email_to_shops.get(em, set())) <= AGENCY_CAP}
    return cleaned


# Stats while we go
already_have = 0
filled = 0
could_not_fill = 0
per_source_filled = defaultdict(int)
high_conf = 0
low_conf = 0

yellow = PatternFill(start_color="FFD6F5D6", end_color="FFD6F5D6", fill_type="solid")  # high conf (green)
orange = PatternFill(start_color="FFFFE6B3", end_color="FFFFE6B3", fill_type="solid")  # low conf (orange)
red = PatternFill(start_color="FFFFC7C7", end_color="FFFFC7C7", fill_type="solid")     # unfilled

for xlsx_row, shop, email_cell, rendered in rows_cache:
    ev = str(rendered).strip() if rendered is not None else ""
    if ev and ev.lower() not in ("#n/a", "n/a", "null"):
        already_have += 1
        continue
    emails = combined_for(shop)
    emails = filter_candidates(shop, emails)
    if not emails:
        could_not_fill += 1
        email_cell.value = ""
        email_cell.fill = red
        continue
    filled += 1
    # Confidence flag: did at least one email domain match the shop brand?
    tokens = set()
    if shop in shop_to_brand:
        tokens.update(brand_tokens(shop_to_brand[shop]))
    tokens.update(shop_slug_tokens(shop))
    for g in shop_groupnames.get(shop, set()):
        tokens.update(brand_tokens(g))
    is_high_conf = any(
        any(tok and tok in em.split("@")[-1].lower() for tok in tokens)
        for em in emails
    )
    if is_high_conf:
        high_conf += 1
        email_cell.fill = yellow
    else:
        low_conf += 1
        email_cell.fill = orange
    for tag in sources_for(shop):
        per_source_filled[tag] += 1
    email_cell.value = "; ".join(sorted(emails))

# Add source-tag + confidence columns so the user can see how each fill came about
ws_master.cell(row=1, column=4).value = "Filled Sources"
ws_master.cell(row=1, column=5).value = "Confidence"
for xlsx_row, shop, email_cell, rendered in rows_cache:
    tags = sources_for(shop)
    ws_master.cell(row=xlsx_row, column=4).value = ",".join(tags) if tags else ""
    ev = str(rendered).strip() if rendered is not None else ""
    already_had = bool(ev) and ev.lower() not in ("#n/a", "n/a", "null")
    # Re-derive confidence for the filled row
    cur_emails = set(str(email_cell.value).split("; ")) if email_cell.value else set()
    cur_emails = {e for e in cur_emails if e}
    if already_had:
        ws_master.cell(row=xlsx_row, column=5).value = "original"
    elif not cur_emails:
        ws_master.cell(row=xlsx_row, column=5).value = "unfilled"
    else:
        tokens = set()
        if shop in shop_to_brand:
            tokens.update(brand_tokens(shop_to_brand[shop]))
        tokens.update(shop_slug_tokens(shop))
        for g in shop_groupnames.get(shop, set()):
            tokens.update(brand_tokens(g))
        is_hc = any(
            any(tok and tok in em.split("@")[-1].lower() for tok in tokens)
            for em in cur_emails
        )
        ws_master.cell(row=xlsx_row, column=5).value = "high" if is_hc else "low"

wb.save(MASTER_XLSX_OUT)

# ── Summary ───────────────────────────────────────────────────────────────
need_fill = filled + could_not_fill
print("==== Enrichment Summary ====")
print(f"Master rows processed:       {len(rows_cache)}")
print(f"  already had emails:        {already_have}")
print(f"  needed filling:            {need_fill}")
print(f"    filled from some source: {filled}")
print(f"      high-confidence (domain-match): {high_conf}")
print(f"      low-confidence  (no domain):    {low_conf}")
print(f"    still missing:           {could_not_fill}")
print()
print("Per-source contributions (rows filled that benefited from source):")
for tag in ["A", "A'", "B", "C", "D", "E", "F", "G"]:
    print(f"  {tag}: {per_source_filled.get(tag, 0)}")
print()
print(f"Output: {MASTER_XLSX_OUT}")
