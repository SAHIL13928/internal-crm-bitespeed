"""Layer in every new mapping source the team uploaded.

New files in data/inputs/ (verified shapes):

  shopurl_master_mapping_clean.csv      1,685 rows — fuller version of the
                                        original shopurl+number+emailids CSV
                                        (we keep using load_shops.py for
                                        contacts; this file is a superset
                                        — re-run that loader to capture
                                        deltas, idempotent).

  Master_filled_v7.xlsx                 multi-sheet:
    'ShopUrl <> Brand Name'             830 rows — brand_name → shopUrl
                                        (source of truth for brand names)
    'Meeting Link <> Brand Name'        5,053 rows — direct meet_link →
                                        shopUrl + attendees + meeting_id
                                        (no regex extraction needed)
    'Fuzzy Brand Name Match'            3,332 rows — meeting_title →
                                        matchedBrand → shopUrl + score
                                        (use only when score is high
                                        enough; we set the bar at 0.8)

  Fireflies Mapping - Sheet1.csv        130,041 rows — every WA message
                                        we've ever associated with a
                                        merchant, with content + shopUrl
                                        + group_name + extracted meet_link.
                                        For mapping purposes we only need
                                        the (group_name, shopUrl) and
                                        (meet_link, shopUrl) pairs.

  Fireflies Mapping - Meeting Link __ Brand Name.csv
                                        same data as the XLSX sheet above
                                        — kept as a CSV alternative;
                                        we read from the XLSX.

What this script does:
  1. Brand name fill from XLSX (only sets when currently NULL)
  2. (group_name, shopUrl) bindings from the 130k WA corpus → bind any
     unbound WhatsAppGroup, add identity-graph binding
  3. Direct (meet_link, shopUrl) bindings from XLSX → bind any orphan
     Meeting + add identity-graph binding
  4. Fuzzy (meeting_title, shopUrl, score) bindings from XLSX → bind
     orphan meetings whose title matches with score >= 0.8

Idempotent throughout. Safe to re-run.

Usage:
    python scripts/enrich_from_new_sources.py --dry-run
    python scripts/enrich_from_new_sources.py
"""
from __future__ import annotations

import argparse
import csv
import logging
import os
import sys
from collections import Counter, defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sqlalchemy.orm import Session  # noqa: E402

from crm_app.db import SessionLocal  # noqa: E402
from crm_app.identity import add_binding  # noqa: E402
from crm_app.models import Meeting, Shop, WhatsAppGroup  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("enrich_from_new_sources")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUTS = os.path.join(ROOT, "data", "inputs")

XLSX_PATH = os.path.join(INPUTS, "Master_filled_v7.xlsx")
WA_CORPUS_CSV = os.path.join(INPUTS, "Fireflies Mapping - Sheet1.csv")

# Score threshold for fuzzy meeting-title binding. 0.8+ is "very confident
# match" per the team's annotated sheet — below that we don't trust it.
FUZZY_SCORE_THRESHOLD = 0.8


# ── 1. brand names ────────────────────────────────────────────────────────
def fill_brand_names(db: Session) -> int:
    """Pull from `ShopUrl <> Brand Name` sheet and fill Shop.brand_name
    where it's currently NULL. We do NOT overwrite manually-edited
    brand names (respect the DB)."""
    if not os.path.exists(XLSX_PATH):
        log.warning("xlsx missing: %s — skipping brand-name fill", XLSX_PATH)
        return 0

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    if "ShopUrl <> Brand Name" not in wb.sheetnames:
        log.warning("sheet 'ShopUrl <> Brand Name' missing")
        return 0

    ws = wb["ShopUrl <> Brand Name"]
    pairs: list[tuple[str, str]] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or len(row) < 2:
            continue
        name, url = row[0], row[1]
        if not url or not name:
            continue
        pairs.append((str(url).strip().lower(), str(name).strip()))

    set_count = 0
    for url, name in pairs:
        shop = db.get(Shop, url)
        if shop is None:
            continue
        if shop.brand_name and shop.brand_name.strip():
            continue
        shop.brand_name = name
        set_count += 1
    log.info("brand names: pairs_in_sheet=%d  shops_filled=%d", len(pairs), set_count)
    return set_count


# ── 2. group_name → shopUrl bindings from the 130k WA corpus ──────────────
def bind_groups_from_wa_corpus(db: Session) -> dict:
    """Each row in the corpus is a real WA message we already attributed
    to a merchant. Extract (group_name, shopUrl) pairs, dedupe by
    most-frequent shop per group_name, and:
      • register an identity-graph binding (group_name → shop_url)
      • set WhatsAppGroup.shop_url for any unbound rows with that name."""
    if not os.path.exists(WA_CORPUS_CSV):
        log.warning("csv missing: %s — skipping WA-corpus pass", WA_CORPUS_CSV)
        return {"pairs": 0, "groups_bound": 0, "graph_bindings": 0}

    pair_counts: dict[tuple[str, str], int] = Counter()
    rows_seen = 0
    with open(WA_CORPUS_CSV, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        header = next(reader, None)  # content, shopUrl, group_name, ...
        for row in reader:
            rows_seen += 1
            if len(row) < 3:
                continue
            shop_url = (row[1] or "").strip().lower()
            group_name = (row[2] or "").strip()
            if not shop_url or not group_name:
                continue
            pair_counts[(group_name, shop_url)] += 1

    # Pick the most-frequent shop per group_name. Collisions (same
    # group_name attributed to multiple shops in the corpus) take the
    # majority winner; if tied, surface as ambiguous (skip).
    by_group: dict[str, dict[str, int]] = defaultdict(dict)
    for (group_name, shop_url), n in pair_counts.items():
        by_group[group_name][shop_url] = n

    groups_bound = ambiguous = 0
    graph_bindings = 0

    for group_name, hits in by_group.items():
        top_count = max(hits.values())
        winners = [s for s, c in hits.items() if c == top_count]
        if len(winners) > 1:
            ambiguous += 1
            continue
        shop_url = winners[0]

        # Update unbound WhatsAppGroup rows with this name.
        for wag in db.query(WhatsAppGroup).filter_by(group_name=group_name).all():
            if wag.shop_url is None:
                wag.shop_url = shop_url
        groups_bound += 1

        # Always-on graph binding (idempotent on natural key).
        try:
            add_binding(
                db,
                "group_name", group_name,
                "shop_url", shop_url,
                source="wa_corpus_authoritative",
                confidence=1.0,  # team-curated source
                evidence_table="data/inputs/Fireflies Mapping - Sheet1.csv",
                evidence_id=group_name,
            )
            graph_bindings += 1
        except ValueError:
            pass

    log.info(
        "wa_corpus: rows=%d distinct_pairs=%d groups_indexed=%d "
        "groups_bound=%d ambiguous=%d graph_bindings=%d",
        rows_seen, len(pair_counts), len(by_group),
        groups_bound, ambiguous, graph_bindings,
    )
    return {
        "pairs": len(pair_counts),
        "groups_bound": groups_bound,
        "graph_bindings": graph_bindings,
        "ambiguous": ambiguous,
    }


# ── 3. direct meet_link → shop_url from XLSX ──────────────────────────────
def bind_meetlinks_direct(db: Session) -> dict:
    """The XLSX has a clean `Meeting Link <> Brand Name` sheet — already
    attributed by the team. Use it to:
      • bind orphan Meeting rows whose meeting_link matches
      • add (meeting_link, shop_url) graph bindings for downstream resolvers"""
    if not os.path.exists(XLSX_PATH):
        return {"pairs": 0, "meetings_bound": 0}

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    if "Meeting Link <> Brand Name" not in wb.sheetnames:
        return {"pairs": 0, "meetings_bound": 0}

    ws = wb["Meeting Link <> Brand Name"]
    # Header: company, date, meeting_title, meeting_link, external_attendees,
    #         internal_attendees, meeting_id, [shopUrl-ish], ...
    # Looking at the sample, column index 7 (0-based) often holds shopUrl
    # but values are often None. We'll trust the row only if it has a
    # meeting_link AND something resembling a shopUrl in any of the
    # trailing columns.
    pairs: dict[str, str] = {}  # meet_link -> shop_url
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row:
            continue
        meeting_link = next((c for c in row if isinstance(c, str)
                             and c.startswith("https://meet.google.com/")), None)
        shop_url = next((c for c in row if isinstance(c, str)
                         and c.endswith(".myshopify.com")), None)
        if not meeting_link or not shop_url:
            continue
        pairs[meeting_link.strip().lower()] = shop_url.strip().lower()

    meetings_bound = graph_bindings = 0
    for link, shop_url in pairs.items():
        # Update unbound Meeting rows whose link matches.
        updated = (
            db.query(Meeting)
            .filter(Meeting.meeting_link.ilike(link), Meeting.shop_url.is_(None))
            .update({"shop_url": shop_url, "mapping_source": "link_xlsx_direct"},
                    synchronize_session=False)
        )
        meetings_bound += updated

        try:
            add_binding(
                db,
                "meeting_link", link,
                "shop_url", shop_url,
                source="xlsx_direct_meet_link",
                confidence=1.0,
                evidence_table="Master_filled_v7.xlsx",
                evidence_id=link,
            )
            graph_bindings += 1
        except ValueError:
            pass

    log.info("meet_link xlsx: pairs=%d meetings_bound=%d graph_bindings=%d",
             len(pairs), meetings_bound, graph_bindings)
    return {"pairs": len(pairs), "meetings_bound": meetings_bound, "graph_bindings": graph_bindings}


# ── 4. fuzzy meeting title → shop ─────────────────────────────────────────
def bind_meetings_via_title(db: Session) -> dict:
    """`Fuzzy Brand Name Match` sheet is the team's hand-curated title→
    shop mapping with confidence scores. We use rows with score >= 0.8
    where shopUrl is supplied, to bind orphan Meeting rows by title."""
    if not os.path.exists(XLSX_PATH):
        return {"used": 0, "meetings_bound": 0}

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    if "Fuzzy Brand Name Match" not in wb.sheetnames:
        return {"used": 0, "meetings_bound": 0}

    ws = wb["Fuzzy Brand Name Match"]
    title_to_shop: dict[str, str] = {}
    used = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or len(row) < 6:
            continue
        _id, title, _matched_brand, shop_url, _tier, score = row[:6]
        if not title or not shop_url or score is None:
            continue
        try:
            sc = float(score)
        except (TypeError, ValueError):
            continue
        if sc < FUZZY_SCORE_THRESHOLD:
            continue
        t = str(title).strip().lower()
        s = str(shop_url).strip().lower()
        if not s.endswith(".myshopify.com"):
            continue
        title_to_shop[t] = s
        used += 1

    meetings_bound = 0
    for title, shop_url in title_to_shop.items():
        updated = (
            db.query(Meeting)
            .filter(Meeting.title.ilike(title), Meeting.shop_url.is_(None))
            .update({"shop_url": shop_url, "mapping_source": "title_fuzzy_xlsx"},
                    synchronize_session=False)
        )
        meetings_bound += updated

    log.info("fuzzy title: rows_used=%d distinct_titles=%d meetings_bound=%d",
             used, len(title_to_shop), meetings_bound)
    return {"used": used, "meetings_bound": meetings_bound}


# ── orchestrator ──────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    db = SessionLocal()
    try:
        log.info("=== 1/4 brand names from XLSX ===")
        fill_brand_names(db)

        log.info("=== 2/4 group_name → shop from WA corpus ===")
        bind_groups_from_wa_corpus(db)

        log.info("=== 3/4 meet_link → shop from XLSX ===")
        bind_meetlinks_direct(db)

        log.info("=== 4/4 fuzzy meeting title → shop ===")
        bind_meetings_via_title(db)

        if args.dry_run:
            db.rollback()
            log.info("dry-run — rolled back")
        else:
            db.commit()
            log.info("committed")
    finally:
        db.close()


if __name__ == "__main__":
    main()
