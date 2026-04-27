"""Populate shops.brand_name from the ENRICHED master spreadsheet.

The 'ShopUrl <> Brand Name' sheet contains ~830 human-curated brand names
keyed by shopUrl. We pull these into shops.brand_name. Only updates rows
where brand_name is currently NULL (manual edits in the DB win).
"""
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from crm_app.db import SessionLocal, engine, Base  # noqa: E402
from crm_app.models import Shop  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(ROOT, "data", "scratch", "Fireflies Mapping (1) - ENRICHED.xlsx")
SHEET = "ShopUrl <> Brand Name"


def load_brand_names():
    Base.metadata.create_all(bind=engine)
    if not os.path.exists(XLSX_PATH):
        print(f"  (skip) {XLSX_PATH} not found")
        return

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        print(f"  (skip) sheet '{SHEET}' missing from workbook")
        return
    ws = wb[SHEET]

    pairs = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or len(row) < 2:
            continue
        name, url = row[0], row[1]
        if not url:
            continue
        url = str(url).strip().lower()
        name = (str(name).strip() if name else None) or None
        if url and name:
            pairs.append((url, name))

    db = SessionLocal()
    set_count = unchanged = missing_shop = 0
    try:
        for url, name in pairs:
            shop = db.get(Shop, url)
            if shop is None:
                missing_shop += 1
                continue
            if shop.brand_name and shop.brand_name.strip():
                unchanged += 1   # respect manual edits
                continue
            shop.brand_name = name
            set_count += 1
        db.commit()
        print(f"brand names processed: {len(pairs)}")
        print(f"  shops updated:       {set_count}")
        print(f"  already set (kept):  {unchanged}")
        print(f"  shop not in DB:      {missing_shop}")
    finally:
        db.close()


if __name__ == "__main__":
    load_brand_names()
